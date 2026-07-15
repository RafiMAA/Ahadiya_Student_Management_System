"""Excel parsing and validation for student import.

Updated format:
  Row 1: Grade | Medium | Academic Year   (class metadata)
  Row 2: Student Name | Gender | DOB (YYYY-MM-DD) | Parent Name | Contact  (headers)
  Row 3+: student data
"""

import re
from datetime import datetime
from typing import IO


def parse_and_validate_excel(file_obj: IO, filename: str) -> dict:
    """Parse an Excel/CSV file and validate each row.
    Returns: { rows: [...], errors: [...], valid_count: int, class_metadata: {...} }
    """
    rows = []
    errors = []
    class_metadata = None

    if filename.endswith('.csv'):
        rows, errors, class_metadata = _parse_csv(file_obj)
    else:
        rows, errors, class_metadata = _parse_xlsx(file_obj)

    valid_count = sum(1 for r in rows if not r.get("errors"))
    return {"rows": rows, "errors": errors, "valid_count": valid_count, "class_metadata": class_metadata}


def _parse_csv(file_obj: IO):
    import csv
    import io
    content = file_obj.read().decode('utf-8')
    reader = csv.reader(io.StringIO(content))
    all_rows = list(reader)

    if not all_rows:
        return [], [{"row": 0, "field": "file", "message": "Empty file"}], None

    headers = [str(v).strip() for v in all_rows[0]]
    rows = []
    errors = []
    
    # Map common header variations
    header_map = {}
    for h in headers:
        hl = h.lower()
        if "name" in hl and "parent" not in hl and ("student" in hl or hl == "full name" or hl == "name"):
            header_map["full_name"] = h
        elif "gender" in hl:
            header_map["gender"] = h
        elif "dob" in hl or "birth" in hl or "date of birth" in hl:
            header_map["dob"] = h
        elif "parent" in hl and "name" in hl:
            header_map["parent_name"] = h
        elif "contact" in hl or "phone" in hl:
            header_map["contact"] = h
        elif "medium" in hl:
            header_map["medium"] = h
        elif "grade" in hl:
            header_map["grade"] = h

    for i, data_row in enumerate(all_rows[1:], start=2):
        raw = {}
        for j, val in enumerate(data_row):
            if j < len(headers):
                raw[headers[j]] = str(val).strip()
                
        mapped = {
            "Student Name": raw.get(header_map.get("full_name", "Student Name"), ""),
            "Gender": raw.get(header_map.get("gender", "Gender"), ""),
            "DOB (YYYY-MM-DD)": raw.get(header_map.get("dob", "DOB (YYYY-MM-DD)"), ""),
            "Parent Name": raw.get(header_map.get("parent_name", "Parent Name"), ""),
            "Contact": raw.get(header_map.get("contact", "Contact"), ""),
            "Medium (Sinhala/Tamil)": raw.get(header_map.get("medium", "Medium (Sinhala/Tamil)"), ""),
            "Grade": raw.get(header_map.get("grade", "Grade"), ""),
        }
        
        row, row_errors = _validate_row(mapped, i)
        rows.append(row)
        errors.extend(row_errors)

    return rows, errors, None


def _parse_xlsx(file_obj: IO):
    from openpyxl import load_workbook
    wb = load_workbook(file_obj, read_only=True)
    ws = wb.active

    # Read all rows first
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        wb.close()
        return [], [{"row": 0, "field": "file", "message": "Empty file"}], None

    headers = [str(v or "").strip() for v in all_rows[0]]

    # Map common header variations
    header_map = {}
    for h in headers:
        hl = h.lower()
        if "name" in hl and "parent" not in hl and ("student" in hl or hl == "full name" or hl == "name"):
            header_map["full_name"] = h
        elif "gender" in hl:
            header_map["gender"] = h
        elif "dob" in hl or "birth" in hl or "date of birth" in hl:
            header_map["dob"] = h
        elif "parent" in hl and "name" in hl:
            header_map["parent_name"] = h
        elif "contact" in hl or "phone" in hl:
            header_map["contact"] = h
        elif "medium" in hl:
            header_map["medium"] = h
        elif "grade" in hl:
            header_map["grade"] = h

    data_start = 0
    rows = []
    errors = []
    for i, excel_row in enumerate(all_rows[data_start + 1:], start=data_start + 2):
        raw = {}
        for j, val in enumerate(excel_row):
            if j < len(headers):
                raw[headers[j]] = str(val) if val is not None else ""

        mapped = {
            "Student Name": raw.get(header_map.get("full_name", "Student Name"), ""),
            "Gender": raw.get(header_map.get("gender", "Gender"), ""),
            "DOB (YYYY-MM-DD)": raw.get(header_map.get("dob", "DOB (YYYY-MM-DD)"), ""),
            "Parent Name": raw.get(header_map.get("parent_name", "Parent Name"), ""),
            "Contact": raw.get(header_map.get("contact", "Contact"), ""),
            "Medium (Sinhala/Tamil)": raw.get(header_map.get("medium", "Medium (Sinhala/Tamil)"), ""),
            "Grade": raw.get(header_map.get("grade", "Grade"), ""),
        }

        row, row_errors = _validate_row(mapped, i)
        rows.append(row)
        errors.extend(row_errors)

    wb.close()
    return rows, errors, None


def _validate_row(raw: dict, row_num: int):
    """Validate a single row and return (parsed_row, errors)."""
    row_errors = []
    full_name = (raw.get("Student Name") or "").strip()
    gender = (raw.get("Gender") or "").strip()
    dob_str = (raw.get("DOB (YYYY-MM-DD)") or "").strip()
    parent_name = (raw.get("Parent Name") or "").strip()
    contact = (raw.get("Contact") or "").strip()
    medium = (raw.get("Medium (Sinhala/Tamil)") or "").strip()
    grade_str = (raw.get("Grade") or "").replace("Grade", "").replace("grade", "").strip()

    if not full_name:
        row_errors.append({"row": row_num, "field": "full_name", "message": "Full name is required"})

    if gender not in ("Male", "Female"):
        row_errors.append({"row": row_num, "field": "gender", "message": f"Invalid gender: '{gender}'. Must be Male or Female"})

    dob = None
    if not dob_str:
        row_errors.append({"row": row_num, "field": "date_of_birth", "message": "Date of birth is required"})
    else:
        try:
            dob = datetime.strptime(dob_str.split(" ")[0], "%Y-%m-%d").date()
        except ValueError:
            row_errors.append({"row": row_num, "field": "date_of_birth", "message": f"Invalid date format: '{dob_str}'. Use YYYY-MM-DD"})

    if not parent_name:
        row_errors.append({"row": row_num, "field": "parent_name", "message": "Parent name is required"})

    if not contact:
        row_errors.append({"row": row_num, "field": "parent_contact", "message": "Contact number is required"})
    elif not re.match(r'^\d{10}$', contact.replace(" ", "")):
        row_errors.append({"row": row_num, "field": "parent_contact", "message": f"Contact must be 10 digits, got: '{contact}'"})

    if medium not in ("Sinhala", "Tamil"):
        row_errors.append({"row": row_num, "field": "medium", "message": f"Invalid medium: '{medium}'. Must be Sinhala or Tamil"})

    grade = None
    if not grade_str:
        row_errors.append({"row": row_num, "field": "grade", "message": "Grade is required"})
    elif not grade_str.isdigit() or not (1 <= int(grade_str) <= 11):
        row_errors.append({"row": row_num, "field": "grade", "message": f"Invalid grade: '{grade_str}'. Must be 1-11"})
    else:
        grade = int(grade_str)

    parsed = {
        "full_name": full_name,
        "gender": gender if gender in ("Male", "Female") else None,
        "date_of_birth": dob,
        "parent_name": parent_name,
        "parent_contact": contact.replace(" ", ""),
        "medium": medium if medium in ("Sinhala", "Tamil") else None,
        "grade": grade,
        "errors": [e["message"] for e in row_errors] if row_errors else [],
    }
    return parsed, row_errors
